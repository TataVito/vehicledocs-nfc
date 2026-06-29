"""
Importador Excel → Supabase
===========================
Uso:
  python importar.py datos.xlsx          # importa vehículos e imágenes
  python importar.py --template          # genera plantilla Excel vacía

Estructura de imágenes:
  imagenes/
    AB1234/
      permiso.jpg
      soap.jpg
      revision.jpg
      gases.jpg
    CD5678/
      permiso.png
      ...

Instalar dependencias:
  pip install supabase openpyxl
"""

import sys
import os
from pathlib import Path

SUPABASE_URL = 'https://tiykdyqpfxnxgvpfvvwm.supabase.co'
SUPABASE_KEY = 'sb_publishable_IEedxSduro5Go1L6mHWLXg_dNxLGrWP'
BUCKET = 'vehicledocs'

COLUMNAS = [
    'patente', 'marca', 'modelo', 'anio', 'color', 'vin',
    'permiso_numero', 'permiso_municipio', 'permiso_venc',
    'soap_compania', 'soap_poliza', 'soap_venc',
    'rt_numero', 'rt_planta', 'rt_venc',
    'rg_numero', 'rg_planta', 'rg_resultado', 'rg_venc',
]

SECCIONES_FOTO = ['permiso', 'soap', 'revision', 'gases']
EXTENSIONES    = ['.jpg', '.jpeg', '.png', '.webp', '.heic']


def crear_template():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print('ERROR: Instala openpyxl con:  pip install openpyxl')
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Vehículos'

    header_fill = PatternFill('solid', fgColor='1565C0')
    header_font = Font(color='FFFFFF', bold=True)

    for col, nombre in enumerate(COLUMNAS, 1):
        cell = ws.cell(row=1, column=col, value=nombre)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(15, len(nombre) + 2)

    # Fila de ejemplo
    ejemplo = ['AB1234', 'Toyota', 'Hilux', '2022', 'Blanco', '9TMATF1J5MT123456',
               'PC-001', 'Municipalidad Santiago', '03/2026',
               'Mapfre', 'POL-789456', '06/2026',
               'RT-123', 'Planta Norte', '08/2026',
               'RG-456', 'Centro Sur', 'Aprobado', '08/2026']
    for col, valor in enumerate(ejemplo, 1):
        ws.cell(row=2, column=col, value=valor)

    nombre = 'plantilla_vehiculos.xlsx'
    wb.save(nombre)
    print(f'✓ Plantilla creada: {nombre}')
    print(f'  Columnas: {", ".join(COLUMNAS)}')
    print(f'\nColoca las imágenes en:')
    print(f'  imagenes/AB1234/permiso.jpg')
    print(f'  imagenes/AB1234/soap.jpg')
    print(f'  imagenes/AB1234/revision.jpg')
    print(f'  imagenes/AB1234/gases.jpg')


def buscar_imagen(patente: str, seccion: str, base: Path):
    """Busca un archivo de imagen para la patente y sección dadas."""
    carpeta = base / patente.upper()
    if not carpeta.exists():
        carpeta = base / patente.lower()
    if not carpeta.exists():
        return None
    for ext in EXTENSIONES:
        ruta = carpeta / f'{seccion}{ext}'
        if ruta.exists():
            return ruta
    return None


def importar(excel_path: str):
    try:
        import openpyxl
    except ImportError:
        print('ERROR: Instala openpyxl con:  pip install openpyxl')
        sys.exit(1)

    try:
        from supabase import create_client
    except ImportError:
        print('ERROR: Instala supabase con:  pip install supabase')
        sys.exit(1)

    print(f'Conectando a Supabase...')
    sb = create_client(SUPABASE_URL, SUPABASE_KEY)

    print(f'Leyendo {excel_path}...')
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Leer encabezados
    headers = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
    col_idx = {h: i for i, h in enumerate(headers)}

    faltantes = [c for c in COLUMNAS if c not in col_idx]
    if faltantes:
        print(f'ADVERTENCIA: Columnas no encontradas en el Excel: {", ".join(faltantes)}')

    carpeta_img = Path(excel_path).parent / 'imagenes'
    if not carpeta_img.exists():
        print(f'INFO: No se encontró carpeta "imagenes/" — se importarán solo los datos de texto.')

    filas = list(ws.iter_rows(min_row=2, values_only=True))
    total = sum(1 for f in filas if f[col_idx.get('patente', 0)])
    print(f'Vehículos a importar: {total}\n')

    ok = 0
    errores = 0

    for i, fila in enumerate(filas, 1):
        patente_val = fila[col_idx['patente']] if 'patente' in col_idx else None
        if not patente_val:
            continue

        patente = str(patente_val).strip().upper()
        print(f'[{i}/{total}] {patente}', end=' ')

        # Construir fila para Supabase
        row = {'patente': patente}
        for col in COLUMNAS[1:]:
            if col in col_idx:
                v = fila[col_idx[col]]
                row[col] = str(v).strip() if v is not None else ''

        # Upsert datos
        res = sb.table('vehiculos').upsert(row, on_conflict='patente').execute()
        if hasattr(res, 'error') and res.error:
            print(f'✗ Error datos: {res.error}')
            errores += 1
            continue

        # Subir imágenes
        fotos_ok = []
        if carpeta_img.exists():
            for sec in SECCIONES_FOTO:
                ruta = buscar_imagen(patente, sec, carpeta_img)
                if not ruta:
                    continue
                storage_path = f'{patente}/{sec}{ruta.suffix.lower()}'
                with open(ruta, 'rb') as f:
                    contenido = f.read()
                mime = {'.jpg':'image/jpeg','.jpeg':'image/jpeg','.png':'image/png',
                        '.webp':'image/webp','.heic':'image/heic'}.get(ruta.suffix.lower(), 'image/jpeg')
                try:
                    sb.storage.from_(BUCKET).upload(storage_path, contenido,
                        {'content-type': mime, 'upsert': 'true'})
                    fotos_ok.append(sec)
                except Exception as e:
                    print(f'\n  ✗ Error foto {sec}: {e}', end='')

        estado = f'✓ datos'
        if fotos_ok:
            estado += f' + fotos: {", ".join(fotos_ok)}'
        print(estado)
        ok += 1

    print(f'\n{"="*40}')
    print(f'Importados: {ok} | Errores: {errores}')
    print(f'URL pública ejemplo: https://tatavito.github.io/vehicledocs-nfc/?p=PATENTE')


if __name__ == '__main__':
    if len(sys.argv) < 2 or sys.argv[1] == '--help':
        print(__doc__)
        sys.exit(0)

    if sys.argv[1] == '--template':
        crear_template()
    else:
        if not os.path.exists(sys.argv[1]):
            print(f'ERROR: No se encuentra el archivo {sys.argv[1]}')
            sys.exit(1)
        importar(sys.argv[1])
